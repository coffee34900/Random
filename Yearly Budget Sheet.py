import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

#Fix widths.
#Still need error handles.
#Make simple tab for drop down.

def create_yearly_budget_openpyxl(filename="YearlyBudget_openpyxl.xlsx"):
    try:
        wb = Workbook()
        
        months = [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ]
        
        instructions_data = [
            ["Welcome to Your Yearly Budget Workbook!"],
            ["Instructions:"],
            ["1. Navigate to each month's sheet to enter your income and expenses."],
            ["2. Use the Dashboard to view summaries for a selected month."],
            ["3. The Yearly Summary sheet aggregates data from all months for an overview."],
            ["4. Ensure all monetary values are entered as numbers."],
            ["5. Charts will update automatically based on your inputs."]
        ]
        if 'Sheet' in wb.sheetnames:
            instr_sheet = wb['Sheet']
            instr_sheet.title = "Instructions"
        else:
            wb.create_sheet("Instructions")
            instr_sheet = wb["Instructions"]
        
        for row_idx, instruction_row in enumerate(instructions_data, start=1):
            cell = instr_sheet.cell(row=row_idx, column=1, value=instruction_row[0])
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 2:
                cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)
        
        bold_font = Font(bold=True, size=12)
        
        header_fill    = PatternFill(start_color="FFFFD700", end_color="FFFFD700", fill_type="solid")  # Gold
        income_fill    = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")  # Light Green
        fixed_fill     = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")  # Light Blue
        variable_fill  = PatternFill(start_color="FFFFB6C1", end_color="FFFFB6C1", fill_type="solid")  # Light Pink
        summary_fill   = PatternFill(start_color="FFFFFFE0", end_color="FFFFFFE0", fill_type="solid")  # Light Yellow
        
        red_fill       = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        green_fill     = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
        
        # Create Monthly Sheets
        for month in months:
            ws = wb.create_sheet(title=month)
            
            # SUMMARY SECTION (Rows 1–6)
            summary_data = [
                ["Monthly Budget for", month],
                ["Total Income:", f"='{month}'!B11"],
                ["Total Fixed Expenses:", f"='{month}'!B21"],
                ["Total Variable Expenses:", f"='{month}'!B30"],
                ["Total Expenses:", f"='{month}'!B21 + '{month}'!B30"],
                ["Leftover/Shortfall:", f"='{month}'!B11 - ('{month}'!B21 + '{month}'!B30)"]
            ]
            for r_idx, row in enumerate(summary_data, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True, size=14)
                    elif r_idx >= 2 and c_idx == 1:
                        cell.font = Font(bold=True, size=12)
                    if r_idx <= 6 and c_idx == 1:
                        cell.fill = summary_fill
                    cell.alignment = Alignment(horizontal="left")
            
            # INCOME SECTION (Rows 8–13)
            income_start_row = 8
            income_data = [
                ["INCOME", "AMOUNT"],
                ["Main Paycheck", 0],
                ["Side Hustle", 0],
                ["Other", 0],
                ["Total Income", "=SUM(B9:B11)"]
            ]
            for r_idx, row in enumerate(income_data, income_start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == income_start_row:  # Header
                        cell.font = bold_font
                        cell.fill = income_fill
                    elif r_idx == income_start_row + len(income_data) - 1:  # "Total Income" row
                        cell.font = bold_font
                        cell.fill = income_fill
                    # Change to ₺:
                    if c_idx == 2 and r_idx != income_start_row:
                        cell.number_format = '"₺"#,##0.00'
            
            # FIXED EXPENSES SECTION (Rows 15–22)
            fixed_start_row = 15
            fixed_data = [
                ["FIXED EXPENSES", ""],
                ["Rent", 0],
                ["Utilities", 0],
                ["Internet/Phone", 0],
                ["Insurance", 0],
                ["Subscriptions", 0],
                ["Loan Payments", 0],
                ["Total Fixed", "=SUM(B16:B21)"]
            ]
            for r_idx, row in enumerate(fixed_data, fixed_start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == fixed_start_row: 
                        cell.font = bold_font
                        cell.fill = fixed_fill
                    elif r_idx == fixed_start_row + len(fixed_data) - 1:  # "Total Fixed"
                        cell.font = bold_font
                        cell.fill = fixed_fill
                    if c_idx == 2 and r_idx != fixed_start_row:
                        cell.number_format = '"₺"#,##0.00'
            
            # VARIABLE EXPENSES SECTION (Rows 23–30)
            variable_start_row = 23
            variable_data = [
                ["VARIABLE EXPENSES", ""],
                ["Groceries", 0],
                ["Transportation", 0],
                ["Dining/Takeout", 0],
                ["Entertainment", 0],
                ["Personal/Health", 0],
                ["Others.", 0],
                ["Total Variable", "=SUM(B24:B29)"]
            ]
            for r_idx, row in enumerate(variable_data, variable_start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == variable_start_row:  
                        cell.font = bold_font
                        cell.fill = variable_fill
                    elif r_idx == variable_start_row + len(variable_data) - 1:  
                        cell.font = bold_font
                        cell.fill = variable_fill
                    if c_idx == 2 and r_idx != variable_start_row:
                        cell.number_format = '"₺"#,##0.00'
            
            column_widths = [25, 15]
            for i, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = width
            
            # Conditional Formatting for Leftover/Shortfall (Cell B6)
            leftover_cell = "B6"
            ws.conditional_formatting.add(
                leftover_cell,
                CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill)
            )
            ws.conditional_formatting.add(
                leftover_cell,
                CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, fill=green_fill)
            )
            
            # Data Validation for Amount Columns
            dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
            ws.add_data_validation(dv)
            
            amount_cells = []
            # Income amounts (B9:B11)
            amount_cells += [f"B{row}" for row in range(9, 12)]
            # Fixed expenses amounts (B16:B21)
            amount_cells += [f"B{row}" for row in range(16, 22)]
            # Variable expenses amounts (B24:B29)
            amount_cells += [f"B{row}" for row in range(24, 30)]
            
            for cell_ref in amount_cells:
                dv.add(cell_ref)
            
            #Pie Chart
            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=24, max_row=29)
            data = Reference(ws, min_col=2, min_row=23, max_row=29)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Variable Expenses Breakdown"
            pie.height = 10 
            pie.width = 10 
            ws.add_chart(pie, "D8")
            
            # Freeze Panes
            ws.freeze_panes = "A7"
        
        #Yearly Summary Sheet
        summary_sheet = wb.create_sheet(title="Yearly Summary")
        
        summary_headers = ["Month", "Total Income", "Total Fixed Expenses", "Total Variable Expenses", "Total Expenses", "Leftover/Shortfall"]
        for col_num, header in enumerate(summary_headers, 1):
            cell = summary_sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for idx, month in enumerate(months, start=2):
            summary_sheet.cell(row=idx, column=1, value=month)
            summary_sheet.cell(row=idx, column=2, value=f"='{month}'!B11")
            summary_sheet.cell(row=idx, column=3, value=f"='{month}'!B21")
            summary_sheet.cell(row=idx, column=4, value=f"='{month}'!B30")
            summary_sheet.cell(row=idx, column=5, value=f"='{month}'!B21 + '{month}'!B30")
            summary_sheet.cell(row=idx, column=6, value=f"='{month}'!B6")
            
            for col in range(2, 7):
                summary_sheet.cell(row=idx, column=col).number_format = '"₺"#,##0.00'
        
        summary_col_widths = [15, 20, 25, 25, 20, 20]
        for i, width in enumerate(summary_col_widths, start=1):
            col_letter = get_column_letter(i)
            summary_sheet.column_dimensions[col_letter].width = width
        
        #Bar Chart for Yearly Summary 
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Yearly Financial Overview"
        bar.y_axis.title = 'Amount (₺)'
        bar.x_axis.title = 'Month'
        
        data = Reference(summary_sheet, min_col=2, min_row=1, max_col=6, max_row=13)
        categories = Reference(summary_sheet, min_col=1, min_row=2, max_row=13)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(categories)
        bar.shape = 4
        summary_sheet.add_chart(bar, "H2")
        
        dashboard = wb.create_sheet(title="Dashboard")
        
        dashboard["A1"] = "Select Month:"
        dashboard["A1"].font = Font(bold=True, size=14)
        dashboard["B1"] = ""  # This will have the drop-down Important
        
        dv_month = DataValidation(type="list", formula1='"' + ','.join(months) + '"', allow_blank=False)
        dashboard.add_data_validation(dv_month)
        dv_month.add("B1")
        dashboard["B1"].alignment = Alignment(horizontal="left")
        dashboard["B1"].font = Font(size=12)
        
        #Selected Month's Summary
        dashboard["A3"] = "Total Income:"
        dashboard["A4"] = "Total Fixed Expenses:"
        dashboard["A5"] = "Total Variable Expenses:"
        dashboard["A6"] = "Total Expenses:"
        dashboard["A7"] = "Leftover/Shortfall:"
        for row in range(3, 8):
            dashboard.cell(row=row, column=1).font = Font(bold=True, size=12)
        
        dashboard["B3"] = "=IF(B1=\"\",\"\",INDIRECT(\"'\"&B1&\"'!B11\"))"
        dashboard["B4"] = "=IF(B1=\"\",\"\",INDIRECT(\"'\"&B1&\"'!B21\"))"
        dashboard["B5"] = "=IF(B1=\"\",\"\",INDIRECT(\"'\"&B1&\"'!B30\"))"
        dashboard["B6"] = "=IF(B1=\"\",\"\",B4+B5)"
        dashboard["B7"] = "=IF(B1=\"\",\"\",B3-B6)"
        
        for row in range(3, 8):
            dashboard.cell(row=row, column=2).number_format = '"₺"#,##0.00'
        
        # Conditional Formatting on Dashboard leftover/shortfall
        dashboard.conditional_formatting.add(
            "B7",
            CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill)
        )
        dashboard.conditional_formatting.add(
            "B7",
            CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, fill=green_fill)
        )
        
        #Pie Chart
        dashboard["A10"] = "Refer to individual month's sheet for detailed charts."
        dashboard["A10"].font = Font(italic=True, size=12)
        
        # Yearly Summary
        dashboard["A12"] = "Yearly Total Income:"
        dashboard["A13"] = "Yearly Total Expenses:"
        dashboard["A14"] = "Yearly Leftover/Shortfall:"
        for row in range(12, 15):
            dashboard.cell(row=row, column=1).font = Font(bold=True, size=12)
        
        dashboard["B12"] = "=SUM('Yearly Summary'!B2:B13)"
        dashboard["B13"] = "=SUM('Yearly Summary'!E2:E13)"
        dashboard["B14"] = "=B12 - B13"
        
        for row in range(12, 15):
            dashboard.cell(row=row, column=2).number_format = '"₺"#,##0.00'
        
        # Conditional Formatting on Yearly leftover/shortfall
        dashboard.conditional_formatting.add(
            "B14",
            CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill)
        )
        dashboard.conditional_formatting.add(
            "B14",
            CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, fill=green_fill)
        )
        
        dashboard["A16"] = "Refer to the Yearly Summary sheet for comprehensive charts and analysis."
        dashboard["A16"].font = Font(italic=True, size=12)
        
        dashboard.column_dimensions['A'].width = 25
        dashboard.column_dimensions['B'].width = 20
        
        wb.save(filename)
        
        print(f"'{filename}' created successfully with multiple sheets, dashboard, and yearly summary!")
    
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    create_yearly_budget_openpyxl("YearlyBudget.xlsx")
